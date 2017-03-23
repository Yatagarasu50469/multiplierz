import openpyxl # Excel 2010 support (.xlsx files)
import xlrd # Older Excel support (.xls files), reading only
import xlwt # Older Excel support (.xls files), writing only
import os
from numbers import Number

from multiplierz.mzReport import ReportReader, ReportWriter, ReportEntry, default_columns



def get_sheet_names(filename):
    """
    Retrieve the names of all sheets in a spreadsheet file.  Works for .xls and .xlsx files.
    """
    
    if filename.lower().endswith('.xls'):
        wb = xlrd.open_workbook(filename, on_demand = True)
        names = wb.sheet_names()
        wb.release_resources()
        return names
    
    elif filename.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        return wb.get_sheet_names()
    
    else:
        raise IOError, "File not a valid Excel format (.xls or .xlsx)."
        


class XLSheetReader(ReportReader):
    def __init__(self, file_name, sheet_name = None, classic_mode = False, autotypecast = True):
        if classic_mode:
            import mzSpreadsheetClassic
            self.__class__ = mzSpreadsheetClassic.XLSheetReader
            return mzSpreadsheetClassic.XLSheetReader.__init__(self, file_name, sheet_name)
        
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.autotypecast = autotypecast
        
        if not os.path.exists(file_name):
            raise IOError, '%s not found!' % file_name
        
        if file_name.lower().endswith('.xls'):
            self.__class__ = XLSReader
            XLSReader.__init__(self, file_name, sheet_name)
        elif file_name.lower().endswith('.xlsx'):
            self.__class__ = XLSXReader
            XLSXReader.__init__(self, file_name, sheet_name)
        else:
            raise IOError, "Invalid extension for file %s given to XLSheetReader!" % file_name


    def __iter__(self):
        raise NotImplementedError
    
    def close(self):
        raise NotImplementedError
    
    def sheet_names(self):
        return get_sheet_names(self.file_name)
    
class XLSheetWriter(ReportWriter):
    def __init__(self, file_name, sheet_name = 'Data',
                 columns = None, default_columns = False, classic_mode = False):
     
        
        if columns:
            self.columns = columns[:]
        elif default_columns:
            self.columns = default_columns + (columns or [])
        else:
            raise IOError, "Columns not specified and default columns not selected!"
    
        if len(self.columns) < len(set([x.lower() for x in self.columns])):
            import defaultdict
            counts = defaultdict(int)
            for col in self.columns:
                counts[col] += 1
                
            raise ValueError, "Column titles appear more than once: %s" % [k for (k, v) in counts.items() if v > 1]
        
        
        if classic_mode:
            import mzSpreadsheetClassic
            self.__class__ = mzSpreadsheetClassic.XLSheetWriter
            mzSpreadsheetClassic.XLSheetWriter.__init__(self, file_name, sheet_name,
                                                        columns, default_columns)   
        elif file_name.lower().endswith('.xls'):
            self.__class__ = XLSWriter
            XLSWriter.__init__(self, file_name, sheet_name, self.columns)
        elif file_name.lower().endswith('.xlsx'):
            self.__class__ = XLSXWriter
            XLSXWriter.__init__(self, file_name, sheet_name, self.columns)
        else:
            raise IOError, "Invalid extension on filename %s given to XLSheetWriter." % file_name
        
        
    

class XLSXReader(XLSheetReader):
    def __init__(self, file_name, sheet_name = None):
        try:
            self.wb = openpyxl.load_workbook(file_name, read_only = True, use_iterators = True)
        except TypeError:
            # I think read_only is the only supported mode in this openpyxl version.
            self.wb = openpyxl.load_workbook(file_name, read_only = True)
        if sheet_name:
            self.sheet = self.wb[sheet_name]
        elif 'Data' in self.wb.get_sheet_names():
            self.sheet = self.wb['Data']
        else:
            sheetname = self.wb.get_sheet_names()[0]
            if len(self.wb.get_sheet_names()) > 1:
                print "Worksheet not specified; defaulting to %s" % sheetname
            self.sheet = self.wb[sheetname]
        
        self.all_sheets = self.wb.get_sheet_names()
        
        try:
            self.columns = [x.value if x.value != None else '' for x in self.sheet.iter_rows().next() if x]
        except StopIteration:
            self.columns = [] # Empty sheet; raise error instead?
            
    def get_row_count(self):
        return self.sheet.max_row
    
    def __iter__(self):
        iterator = self.sheet.iter_rows()
        iterator.next() # Skip header.
        for row in iterator:
            values = [x.value if x.value != None else '' for x in row]
            if not any(values): continue
            yield ReportEntry(self.columns, values, self.autotypecast)
    
    def close(self):
        self.wb._archive.close() # I have bad words for the openpyxl writers.
        try:
            del self.wb
            del self.sheet
        except AttributeError:
            print "WARNING: Close called multiple times. (XLSXReader.)"
        
        
class XLSXWriter(XLSheetReader):
    def __init__(self, file_name, sheet_name = 'Data', columns = None):
        if not columns: raise IOError, "No columns!  Also, use XLSheetReader instead of XLSXWriter directly."
        
        self.file_name = file_name
        self.sheet_name = sheet_name
        
        if os.path.exists(file_name):
            self.wb = openpyxl.load_workbook(file_name)
            
            try:
                self.wb.remove_sheet(self.wb.get_sheet_by_name(sheet_name))
            except (ValueError, KeyError):
                pass
            
            self.sheet = self.wb.create_sheet(title = sheet_name)
        else:
            self.wb = openpyxl.Workbook()
            self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
            self.sheet = self.wb.create_sheet(title = sheet_name)
        
        # Electing to not try and get column order from pre-existing sheet.  Too many problems.
        self.currentRow = 1
        self.write(self.columns)
    
    def write(self, row, metadata = None):
        if metadata:
            raise NotImplementedError, "Non-comtypes Excel interface can't handle Excel metadata."
        
        if len(row) > len(self.columns):
            missing = set([str(x).lower() for x in row.keys()]) - set([str(x).lower() for x in self.columns])
            raise ValueError, "Row has extra columns: %s" % missing 
        # It should be impossible for this error to show an empty set!
        
        if isinstance(row, dict):
            row = dict(row)
            for key in row.keys(): # There's probably a better way to do that.
                lKey = key
                if lKey != key and (key in row): row[lKey] = row[key]
            try:
                row = [row[x] for x in self.columns]
            except KeyError as err:
                print (sorted(row.keys()), '\n', sorted(self.columns))
                raise err
        
        for index, value in enumerate(row, start = 1):
            cell = self.sheet.cell(row = self.currentRow, column = index)
            if value == None:
                cell.value = ''
            elif isinstance(value, Number):
                cell.value = value
            else:
                cell.value = str(value)
        self.currentRow += 1
    
    def close(self):
        try:
            self.wb.save(self.file_name)
            print "Closed %s (Sheet %s)" % (self.file_name, self.sheet_name)
        except IOError as err: 
            # Goodness knows how much data has been lost the grabby
            # lock Excel puts on open files, so we try to at least save
            # things somewhere.
            print "Overwrite failed..."
            self.wb.save(self.file_name[:-5] + '.OVERWRITE.xlsx')
            print "Closed %s (Sheet %s)" % (self.file_name[:-5] + '.OVERWRITE.xlsx',
                                            self.sheet_name)
        
        
        
class XLSWriter(XLSheetWriter):
    def __init__(self, file_name, sheet_name = 'Data', columns = None):
        if not columns: raise IOError, "No columns!  Also, use XLSheetReader instead of XLSWriter directly."
        
        self.file_name = file_name
        self.sheet_name = sheet_name
        
        self.previous_data = []
        # *Could probably* use xlutils to modify .xlses more effeciently.  But instead
        # will have the object carry the old file data while it exists.  The file can't
        # be too large, right?
        if os.path.exists(file_name):
            oldWB = xlrd.open_workbook(file_name)
            for oldSheetName in [s for s in oldWB.sheet_names() if s != sheet_name]:
                oldSheet = oldWB.sheet_by_name(oldSheetName)
                #self.previous_data = (oldSheetName, list(oldSheet.iter_rows()))
                self.previous_data.append((oldSheetName, [oldSheet.row(x) for x in range(0, oldSheet.nrows)]))
        
        self.wb = xlwt.Workbook()
        self.sheet = self.wb.add_sheet(sheet_name)
        self.columns = columns
        
        self.currentRow = 0
        self.write(self.columns)
        
    def write(self, row, metadata = None):
        if metadata:
            raise NotImplementedError, "Non-comtypes Excel interface can't handle Excel metadata."
        
        if len(row) > len(self.columns):
            raise ValueError, "Row is missing values for some columns."
        elif len(row) < len(self.columns):
            raise ValueError, "Row is too long for sheet columns."
        
        if isinstance(row, dict):
            row = dict(row)
            row = [row[x] for x in self.columns]         
    
        for index, value in enumerate(row):
            self.sheet.write(self.currentRow, index, value if (isinstance(value, Number) or value == None) else str(value))
        self.currentRow += 1
        
    def close(self):
        for previousSheetName, previousSheetRows in self.previous_data:
            previousSheet = self.wb.add_sheet(previousSheetName)
            for rowNum, row in enumerate(previousSheetRows):
                for colNum, value in enumerate(row):
                    previousSheet.write(rowNum, colNum, value.value)
        
        try:
            flptr = open(self.file_name, 'w+b')
            self.wb.save(flptr)
            flptr.close()
            print "Closed %s (Sheet %s)" % (self.file_name, self.sheet_name)
        except IOError as err:
            self.wb.save(self.file_name[:-4] + '.~OVERWRITE.xls')
            print "Couldn't overwrite %s" % self.sheet_name
            print "Closed %s (Sheet %s)" % (self.file_name[:-4] + '.~OVERWRITE.xls',
                                            self.sheet_name)

        
class XLSReader(XLSheetWriter):
    def __init__(self, file_name, sheet_name = None):
        self.wb = xlrd.open_workbook(file_name, on_demand = True)
        
        if sheet_name:
            try:
                self.sheet = self.wb.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                raise IOError, "%s has no sheet named %s" % (file_name, sheet_name)
        elif 'Data' in self.wb.sheet_names():
            self.sheet = self.wb.sheet_by_name('Data')
        else:
            sheetname = self.wb.sheet_names()[0]
            print "Warning:  Worksheet name not specified; defaulting to %s" % sheetname
            self.sheet = self.wb.sheet_by_name(sheetname)
            
        self.all_sheets = self.wb.sheet_names()[0]
                
        self.columns = [x.value if x.value != None else '' for x in self.sheet.row(0)]
    
    def __iter__(self):
        for index in range(1, self.sheet.nrows):
            row = self.sheet.row(index)
            yield ReportEntry(self.columns, [x.value for x in row], self.autotypecast)
    
    def close(self):
        self.wb.release_resources()
        

        







